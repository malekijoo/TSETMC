import os
import getpass
import pandas as pd
from openpyxl import load_workbook
from pandas import ExcelWriter

if 'posix' in os.name:
    path = '/Users/amir/Documents/CODE/Python/start/'
elif 'nt' in os.name:
    path = "C:\\Users\\{}\\Desktop\\start\\".format(str(getpass.getuser()))


def write_to_execl(obj):
    dic = {
        'name': obj.name,
        'price': obj.price,
        'ydp': obj.ydp,
        'wp_fi': obj.wp_fi,
        'change': obj.change,
        'change_per': obj.change_per,
        'd_high': obj.d_high,
        'd_low': obj.d_low,
        'eps': obj.eps,
        'pe': obj.pe,
        # 'whole_volume': obj.whole_volume,
        'trade_volume': obj.trade_volume,
        'trade_value': obj.trade_value,
        'trade_amount': obj.trade_amount,

        'tepix': obj.tepix,
        'dollar': obj.dollar,
        'purchase_q': obj.purchase_q,
        'sale_q ': obj.sale_q,
        'condition': obj.condition,
        'datetime': obj.datetime
    }
    filename = 'dataframe.xlsx'
    sheetname = 'sheet1'
    data = pd.DataFrame(dic, index=[0])
    address = path + '{}'.format(filename)

    if not os.path.exists(address):
        print(' "{}" will be created in {}'.format(filename, address))
        with pd.ExcelWriter(address, engine='openpyxl') as writer:
            data.to_excel(writer, sheet_name=sheetname, index=None)
            writer.save()

    elif os.path.exists(address):
        writer: ExcelWriter
        with pd.ExcelWriter(address, engine='openpyxl') as writer:
            book = load_workbook(address)
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            start_row = writer.sheets[sheetname].max_row
            data.to_excel(writer, sheet_name=sheetname, startrow=start_row, index=None, header=False)
            writer.save()
